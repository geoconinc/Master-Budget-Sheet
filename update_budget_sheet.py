# All the imports
import pandas as pd
import os
import time
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from office365.sharepoint.client_context import ClientContext
from office365.runtime.auth.user_credential import UserCredential

TARGET_DIR = r"C:\GeoconBST_Summary"
EXE_NAME = "update_budget_sheet.exe"

# Credentials from environment (do not hardcode secrets)
EMAIL = os.environ.get("SHAREPOINT_EMAIL", "")
PASSWORD = os.environ.get("SHAREPOINT_PASSWORD", "")

SHAREPOINT_SITE = "https://geoconmail.sharepoint.com/sites/GeoconCentral"
SHAREPOINT_DOC_LIB = "Shared Documents/GeoconDocuments/ProjectNumbers"
MASTER_FILENAME = "all_geocon_budgets.xlsx"

script_dir = os.path.dirname(os.path.abspath(sys.argv[0]))
LOCAL_PATH = os.path.join(script_dir, MASTER_FILENAME)

MASTER_PATH = MASTER_FILENAME

# Sanity checks for authentication
if not EMAIL or not PASSWORD:
    raise EnvironmentError(
        "SHAREPOINT_EMAIL or SHAREPOINT_PASSWORD not set in environment variables"
    )

if not MASTER_FILENAME.lower().endswith(".xlsx"):
    raise ValueError("Only .xlsx files are supported")

# Set up a basic connection to the client using the OFFICE-365-REST-API
ctx = ClientContext(SHAREPOINT_SITE).with_credentials(
    UserCredential(EMAIL, PASSWORD)
)


# Downloading the global sharepoint file
def download_from_sharepoint():
    print("Starting download from SharePoint...")
    try:
        file_url = f"/sites/GeoconCentral/{SHAREPOINT_DOC_LIB}/{MASTER_FILENAME}".replace(" ", "%20")
        file = ctx.web.get_file_by_server_relative_url(file_url)
        with open(LOCAL_PATH, "wb") as f:
            file.download(f).execute_query()
        print("Download complete")

    # Print exception (in case)
    except Exception as e:
        print(f"Failed to download file: {e}")
        raise


# Upload the new summary file back to the sharepoint
def upload_to_sharepoint():
    print("Uploading file to SharePoint...")
    try:
        target_folder = ctx.web.get_folder_by_server_relative_url(SHAREPOINT_DOC_LIB)
        with open(LOCAL_PATH, "rb") as f:
            target_folder.upload_file(MASTER_FILENAME, f).execute_query()
        print("Upload complete")
    except Exception as e:
        print(f"Failed to upload file: {e}")
        raise


def load_source_excel():
    path = r"C:\GeoconBST_Summary\Project Summary.xlsx"
    if not os.path.exists(path):
        print(f"The source Excel file could not be found at:\n{path}")
        exit()
    print(f"Loading file: {os.path.basename(path)}")
    return pd.read_excel(path)


def prepare_rows(df: pd.DataFrame):
    df = df.copy()

    # Fill in missing expected columns with 0s
    if "Revenue % Complete" not in df.columns:
        df["Revenue % Complete"] = 0
    if "Backlog" not in df.columns:
        df["Backlog"] = 0
    if "All Effort" not in df.columns:
        df["All Effort"] = 0
    if "Receivable Balance" not in df.columns:
        df["Receivable Balance"] = 0
    if "TD Budget Effort" not in df.columns:
        df["TD Budget Effort"] = 0

    # Clean and filter
    df = df[~df["Project Code"].astype(str).str.startswith("*")]
    df = df[df["Project Code"].notna() & (df["Project Code"].astype(str).str.strip() != "")]

    return df[
        [
            "Revenue % Complete",         # 0
            "Backlog",                    # 1
            "Project Name",               # 2
            "Project Code",               # 3
            # "All Effort",              # removed
            # "Receivable Balance",      # removed
            "Project Manager Name",       # 4
            "Project Client Name",        # 5
            "Project Organization Name",  # 6
            "Project Director Name",      # 7
            "TD Budget Effort"            # 8
        ]
    ].values.tolist()


def update_master(rows):
    global MASTER_PATH
    MASTER_PATH = LOCAL_PATH  # ensure both saving and upload use the same file

    if os.path.exists(MASTER_PATH):
        wb = load_workbook(MASTER_PATH)
        ws = wb.active

        if ws.max_row > 1:
            for row in range(2, ws.max_row + 1):
                for col in range(1, 15):
                    ws.cell(row=row, column=col).value = None
                    ws.cell(row=row, column=col).fill = PatternFill(fill_type=None)
        else:
            wb = Workbook()
            ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active

    headers = [
        "Project Number",   # A
        "% Complete",       # B
        "$ Left",           # C
        "Total Budget",     # D
        "Project Name",     # E
        "Client",           # F
        "Project Manager",  # G
        "Project Director", # H
        "Organization"      # I
    ]

    col_map = [
        (3, 1),  # Project Code        → A (used internally)
        (0, 2),  # % Complete          → B
        (1, 3),  # $ Left              → C
        (8, 4),  # Total Budget        → D
        (2, 5),  # Project Name        → E
        (5, 6),  # Client              → F
        (4, 7),  # Project Manager     → G
        (7, 8),  # Project Director    → H
        (6, 9)   # Organization        → I
    ]

    header_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")

    # Horizontal legend in row 1
    legend_labels = [
        ("% Complete", None),
        ("50%–74%", "FAFAD2"),
        ("75%–89%", "FFE4B5"),
        ("90%–100%", "F08080"),
        (">100%", "D8BFD8")
    ]

    for col_index, (text, color) in enumerate(legend_labels, start=1):
        cell = ws.cell(row=1, column=col_index, value=text)
        if color:
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")

    # Main column headers in row 2
    for col, title in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=title)
        cell.fill = header_fill

    for i, row_data in enumerate(rows, start=3):
        for idx, col in col_map:
            value = row_data[idx]
            cell = ws.cell(row=i, column=col)
            if col == 2 and isinstance(value, (int, float)):
                rounded_value = round(value)
                cell.value = rounded_value
                if 50 <= rounded_value < 75:
                    cell.fill = PatternFill(start_color="FAFAD2", end_color="FAFAD2", fill_type="solid")
                elif 75 <= rounded_value < 90:
                    cell.fill = PatternFill(start_color="FFE4B5", end_color="FFE4B5", fill_type="solid")
                elif 90 <= rounded_value <= 100:
                    cell.fill = PatternFill(start_color="F08080", end_color="F08080", fill_type="solid")
                elif rounded_value > 100:
                    cell.fill = PatternFill(start_color="D8BFD8", end_color="D8BFD8", fill_type="solid")
            elif col == 3 and isinstance(value, (int, float)):
                cell.value = f"${value:,.2f}"
            elif col == 4 and isinstance(value, (int, float)):
                cell.value = f"${value:,.2f}"
            else:
                string_cols = [1, 5, 6, 7, 8, 9]
                if col in string_cols and pd.notna(value):
                    cell.value = str(value).strip()
                    cell.number_format = "@"
                else:
                    cell.value = value

    widths = {
        'A': 15, 'B': 15, 'C': 15, 'D': 18,
        'E': 30, 'F': 20, 'G': 25,
        'H': 25, 'I': 25, 'K': 22
    }
    for letter, width in widths.items():
        ws.column_dimensions[letter].width = width

    ws.auto_filter.ref = f"A2:J{len(rows) + 2}"
    ws.freeze_panes = "A3"

    try:
        wb.save(MASTER_PATH)
        print(f"Workbook updated with {len(rows)} entries")
    except Exception as e:
        print(f"Failed to save workbook: {e}")
        raise


def run_update_process():
    try:
        start_time = time.time()
        print("Connecting to SharePoint...")
        download_from_sharepoint()
        print("Processing local data...")
        df = load_source_excel()
        rows = sorted(prepare_rows(df), key=lambda x: str(x[3]))
        print("Rebuilding Excel file...")
        update_master(rows)
        upload_to_sharepoint()

        # Clean up local file
        if os.path.exists(LOCAL_PATH):
            try:
                os.remove(LOCAL_PATH)
                print(f"Deleted temporary file: {LOCAL_PATH}")
            except Exception as e:
                print(f"Failed to delete temporary file: {e}")

        end_time = time.time()
        print(f"Process complete in {round(end_time - start_time, 2)} seconds")
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise


if __name__ == "__main__":
    run_update_process()
